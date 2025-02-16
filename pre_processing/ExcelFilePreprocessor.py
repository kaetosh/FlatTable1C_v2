# -*- coding: utf-8 -*-
"""
Created on Mon Dec 16 16:35:32 2024

@author: kaetosh

После пересохранения файлов в актуальный формат .xlsx перед их загрузкой в pandas
они должны пройти предварительную обработку, а именно
- снятие объединения ячеек
- добавления столбца с номерами группировок строк (используется для создания плоской таблицы)
- добавление столбца с признаком курсивного шрифта (актуально для анализа счета в УПП, строки с курсивом
это промежуточные итоги, для исключения в сводном файле)
"""
from typing import List
import openpyxl
from pathlib import Path
from config import analysis_fields
from additional.progress_bar import progress_bar

class ExcelFilePreprocessor:
    @staticmethod
    def preprocessor_openpyxl(excel_files: List[Path]) -> None:
        for i, oFile in enumerate(excel_files):
            progress_bar(i + 1, len(excel_files), prefix='Предобработка исходных файлов:')
            workbook = None
            try:
                workbook = openpyxl.load_workbook(oFile)
            except Exception as e:
                print(f'''{oFile}: Ошибка обработки файла.
                Возможно открыт обрабатываемый файл.
                Закройте этот файл и снова запустите скрипт.
                Ошибка: {e}''')
            sheet = workbook.active
            # Снимаем объединение ячеек
            merged_cells_ranges = list(sheet.merged_cells.ranges)
            for merged_cell_range in merged_cells_ranges:
                sheet.unmerge_cells(str(merged_cell_range))
            # Столбец с уровнями группировок
            sheet.insert_cols(idx=1)
            for row_index in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_index, column=1)
                cell.value = sheet.row_dimensions[row_index].outline_level
            # Столбец с признаком курсива
            sheet.insert_cols(idx=2)
            for row in sheet.iter_rows(values_only=True):
                found_value = next((value for value in [analysis_fields.upp.version_1c_id,
                                                        analysis_fields.notupp.version_1c_id] if value in row), None)
                if found_value is not None:
                    kor_schet_col_index = row.index(found_value) + 1  # Мы добавляем 1, потому что индексация начинается с 0
                    for row_index in range(2, sheet.max_row + 1):  # Мы начинаем с цифры 2, чтобы пропустить название
                        kor_schet_cell = sheet.cell(row=row_index, column=kor_schet_col_index)
                        new_cell = sheet.cell(row=row_index, column=2)
                        new_cell.value = 1 if kor_schet_cell.font and kor_schet_cell.font.italic else 0
                    break
            workbook.save(oFile)
            workbook.close()

